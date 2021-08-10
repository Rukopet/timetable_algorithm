import pandas as pd
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from timetable_Genetic_Algorithm.utils.algorithm_settings import AlgorithmSettings
from timetable_Genetic_Algorithm.utils.constants import DAYS_OF_WEEK_RUSSIAN


def util_diff_tuples(x):
    return int(x[0]) - int(x[1])


def print_groups_in_excel(groups: list, row_begin: int, columns_begin: int, ws: Worksheet) -> None:
    for index, group in enumerate(groups):
        ws.cell(row_begin, columns_begin + index).value = f'{group[0]}{group[1]}'


class Individ:
    """ individ with reload __str__ """

    def __init__(self, dict_ind: dict, settings: AlgorithmSettings):
        self.dict_individ = dict_ind
        self.settings = settings

    def into_excel_file(self, path: str = "", file_name: str = "default.xls"):
        wb = openpyxl.Workbook()
        ws = wb.active

        lessons_in_day = self.settings.AMOUNT_TIMELINES_IN_DAY

        ws.cell(1, 1).value = "Дни недели"
        ws.cell(2, 1).value = "Классы"
        ws.cell(2, 2).value = "Классы и предметы"

        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=35)
        print_groups_in_excel(self.settings.GROUPS_LIST, 2, 2, ws)
        for i in range(6 if self.settings.bool_SCHOOL_STUDY_SATURDAY else 5):
            day = DAYS_OF_WEEK_RUSSIAN[i]
            ws.merge_cells(start_row=i * lessons_in_day + 3, start_column=1,
                           end_row=(i + 1) * lessons_in_day + 2, end_column=1)
            ws.cell(i * lessons_in_day + 3, 1).value = day
        for index_x, (key, value) in enumerate(self.dict_individ.items()):
            x = index_x + 3
            for index_y, (k, val) in enumerate(value.items()):
                y = index_y + 2
        wb.save(path + file_name)

    class MarkdownGroupsInExcel:
        """
        util class for Individ
        if u have x and y take dict from self.settings.GROUPS_LIST nad coordinates
        """

        def __init__(self, settings: AlgorithmSettings, row_begin: int, row_end: int, column_begin: int):
            self.settings = settings
            self.row_begin = row_begin
            self.row_end = row_end
            self.columns_begin = column_begin
            self.group_lesson = {}

        def adding_lesson_to_group(self, timeline_begin: int, timeline_end: int, dict_lessons: dict):
            for timeline in range(timeline_begin, timeline_end):
                timeline_lessons = dict_lessons.get(timeline)
                for audience, lesson in timeline_lessons.items():
                    if lesson is None:
                        continue
                    tmp_list = [lesson[i] for i in range(1, len(lesson))]
                    tmp_list.append(audience)
                    self.group_lesson[lesson[0]] = self.group_lesson.get(lesson[0], []) \
                        .append(tuple(tmp_list))

        def __take_value_from_lesson_tuple(self, lesson: tuple or None):
            if self.settings.DEBUG == 1:
                return "-" if lesson is None else f''
            # if lesson

        def print_into_excel_groups_columns(self, ws: Worksheet):
            for y, group in enumerate(self.settings.GROUPS_LIST):
                for x in range(self.row_begin + 1, self.row_end + 1):
                    # val =
                    ws.cell(x, y)

    #     """
    #      util class for Individ
    #      if u have x and y take dict from self.settings.GROUPS_LIST nad coordinates
    #     """
    #
    #     def __init__(self, settings: AlgorithmSettings, x: int, y: int):


class Lesson:
    """ intersection of main many; timeline unit """

    def __init__(self, num_groups: int, letter_group: str, pedagog_name: str):
        self.audience = None
        self.linked = False
        self.locked = False
        self.pedagog_name = pedagog_name
        self.current_tuple = None


class GeneratorLessons:
    """ main generator whole table, one on which individ """
    j = 1

    def __init__(self, settings: AlgorithmSettings, week: list):
        self._settings = settings
        self.main_tuple_list = week
        # self.main_dataDF = pd.DataFrame.from_dict(settings.main_data, )
        # print(self.main_dataDF)

    def gen_lesson(self):
        pass

    def gen_week(self):
        pass

    @staticmethod
    def gen_overall_pool(settings: AlgorithmSettings):
        df = settings.data_from_front.groupsJSON.valueDF
        return [tuple([tuple([group[0], group[1]]),
                       key,
                       value.get("pedagog"),
                       util_diff_tuples(df.loc[df["number"] == group[0]][["max_days", "saturday_not_study"]].values[0]),
                       settings.TIME_FIRST_LESSON_SECOND_SHIFT
                       if df.loc[df["number"] == group[0]]["second_shift_study"].values[0] else 0])
                for group in settings.GROUPS_LIST
                for key, value in settings.get_group_data(group).items()
                for _ in range(value["load"])]
