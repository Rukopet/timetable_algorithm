from typing import List, Generator, Tuple, Dict, Optional
from timetable_Genetic_Algorithm.utils.our_typing import Group

from timetable_Genetic_Algorithm.utils.algorithm_settings import AlgorithmSettings

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from timetable_Genetic_Algorithm.utils.constants import DAYS_OF_WEEK_RUSSIAN
from copy import deepcopy
import logging


def util_diff_tuples(x):
    return int(x[0]) - int(x[1])


def print_groups_in_excel(groups: list, row_begin: int, columns_begin: int, ws: Worksheet) -> None:
    for index, group in enumerate(groups):
        ws.cell(row_begin, columns_begin + index).value = f'{group[0]}{group[1]}'


class Individ:
    """ individ with reload __str__ """

    """ amount blanks between days of week """
    moderation_amount_blank_rows_between_day_of_week = 0

    # TODO check this value ->, needed more objective, less heuristic
    moderation_absolute_maximum_rows = 500

    moderation_ident_from_left = 0

    def __init__(self, dict_ind: dict, settings: AlgorithmSettings):
        self.dict_individ = dict_ind
        self.settings = settings

    def into_excel_file(self, path: str = "", file_name: str = "default.xlsx", axis: bool = False):

        wb = openpyxl.Workbook()
        ws = wb.active

        Individ.__print_markdown_into_excel(ws)
        self.__merge_cells(ws)

        """
        key -> tuple(group): value -> generator(tuple(x, y))
        u can call next for use this generators
        """
        dict_of_generators = self.__get_dict_of_generators()

        print_groups_in_excel(self.settings.GROUPS_LIST, 2, 2, ws)
        for key, value in self.dict_individ.items():
            whole_groups = deepcopy(self.settings.GROUPS_LIST)
            output = {}
            for k, val in value.items():
                try:
                    whole_groups.remove(val[0])
                except ValueError:
                    logging.debug(f"in excel print {val[0]}")
                except Exception as e:
                    logging.debug(f"in excel print {val[0]}")
                output[val[0]] = self.__get_str_from_tuple(val, output.get(val[0]))
            Individ.__print_out(output, whole_groups, ws, dict_of_generators)
        wb.save(path + file_name)

    def __get_str_from_tuple(self, lesson_tuple: tuple, current_value: Optional[str]) -> str:
        if self.settings.DEBUG == 0:
            return f'{current_value}\n{lesson_tuple[2]}' if current_value else lesson_tuple[2]
        elif self.settings.DEBUG == 1:
            return current_value + ' ' + ', '.join(lesson_tuple) if current_value else ', '.join(lesson_tuple)

    @staticmethod
    def __print_out(out_dict: Dict[Group, str], groups: List[Group], ws: Worksheet,
                    generator_dict: Dict[Group, Generator[Tuple[int, int], None, None]]) -> None:
        for key, value in out_dict.items():
            current_gen = generator_dict.get(key)
            x, y = next(current_gen)
            ws.cell(x, y).value = value
        for empty_group in groups:
            current_gen = generator_dict.get(empty_group)
            x, y = next(current_gen)
            ws.cell(x, y).value = "--"


    @staticmethod
    def __print_markdown_into_excel(ws: Worksheet) -> None:
        ws.cell(1, 1).value = "Дни недели"
        ws.cell(2, 1).value = "Классы"
        ws.cell(1, 2).value = "Классы и предметы"

    def __merge_cells(self, ws: Worksheet) -> None:
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1 + self.settings.GROUPS_LIST.__len__())

        lessons_in_day = self.settings.AMOUNT_TIMELINES_IN_DAY

        for i in range(6 if self.settings.bool_SCHOOL_STUDY_SATURDAY else 5):
            day = DAYS_OF_WEEK_RUSSIAN[i]

            """ merging cells for day of week in first row """
            ws.merge_cells(start_row=i * lessons_in_day + 3, start_column=1,
                           end_row=(i + 1) * lessons_in_day + 2, end_column=1)
            ws.cell(i * lessons_in_day + 3, 1).value = day

    def __get_dict_of_generators(self) -> Dict[Group, Generator[Tuple[int, int], None, None]]:
        """ gen dict generators for useful getting next groups coordinate in excel table """

        return {
            group: self.__get_generator_for_group(*self.__get_group_first_coordinates(group))
            for group in self.settings.GROUPS_LIST
        }

    def __get_group_first_coordinates(self, *args: [int, str], axis: bool = False) -> Tuple[int, int]:
        """ return main coordinates of group """

        if not axis:
            try:
                index = self.settings.GROUPS_LIST.index(tuple(*args))
                return 2, 2 + self.moderation_amount_blank_rows_between_day_of_week + index
            except Exception as e:
                logging.debug(f"need check group {args}", exc_info=e)
                return 0, 0
        else:
            try:
                index = self.settings.GROUPS_LIST.index(args)
                return 2 + self.moderation_amount_blank_rows_between_day_of_week + index, 2
            except Exception as e:
                logging.debug(f"need check group in 2 {args}", exc_info=e)
                return 0, 0

    # TODO axis not usable
    # TODO fix Generator annotations
    def __get_generator_for_group(self, *args: int,
                                  axis: bool = False) -> Generator[Tuple[int, int], None, None]:
        blank_val = 0
        begin_row = args[0]
        for row in range(args[0], self.moderation_absolute_maximum_rows):
            yield row + blank_val, args[1]
            if (row - begin_row) % self.settings.AMOUNT_TIMELINES_IN_DAY == 0\
                    and self.moderation_amount_blank_rows_between_day_of_week > 0:
                blank_val += 1
